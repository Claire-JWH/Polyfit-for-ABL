import json


def main(NUM):
    Jlabel = ['J37', 'J36', 'J35', 'J34']
    Tlabel = ['T37', 'T36', 'T35', 'T34']
    Rlabel = ['R37', 'R36', 'R35', 'R34']
    Clabel = ['C37', 'C36', 'C35', 'C34']
    Glabel = ['G3334', 'G3433', 'G3435', 'G3534', 'G3536', 'G3635', 'G3637', 'G3736']
    Olabel = ['O3334', 'O3433', 'O3435', 'O3534', 'O3536', 'O3635', 'O3637', 'O3736']

    Xj, Yj = [], []
    Xt, Yt = [], []
    Xr, Yr = [], []
    Xc, Yc = [], []
    Xg, Yg = [], []
    Xo, Yo = [], []

    # 打开文件,r是读取,encoding是指定编码格式
    with open('../json/' + NUM + '.json', 'r', encoding='utf-8') as fp:
        print(type(fp))  # 输出结果是 <class '_io.TextIOWrapper'> 一个文件类对象

        # load()函数将fp(一个支持.read()的文件类对象，包含一个JSON文档)反序列化为一个Python对象
        data = json.load(fp)

        print(type(data))  # 输出结果是 <class 'dict'> 一个python对象,json模块会根据文件类对象自动转为最符合的数据类型,所以这里是dict
        fp.close()

    for i in data['shapes']:
        if i['label'][0:3] in Jlabel and i['label'][2:4] != '7D':
            Xj.append(i['points'][0][0])
            Yj.append(i['points'][0][1])
        elif i['label'][0:3] in Tlabel:
            Xt.append(i['points'][0][0])
            Yt.append(i['points'][0][1])
        elif i['label'][0:3] in Rlabel:
            Xr.append(i['points'][0][0])
            Yr.append(i['points'][0][1])
        elif i['label'][0:3] in Clabel and i['label'][2:4] != '7D':
            Xc.append(i['points'][0][0])
            Yc.append(i['points'][0][1])
        elif i['label'] in Glabel:
            Xg.append(i['points'][0][0])
            Yg.append(i['points'][0][1])
        elif i['label'] in Olabel:
            Xo.append(i['points'][0][0])
            Yo.append(i['points'][0][1])

    print("NUM:", NUM)
    print("Xj:", Xj)
    print("Yj:", Yj)
    print("Xt:", Xt)
    print("Yt:", Yt)
    print("Xr:", Xr)
    print("Yr:", Yr)
    print("Xc:", Xc)
    print("Yc:", Yc)
    print("Xg:", Xg)
    print("Yg:", Yg)
    print("Xo:", Xo)
    print("Yo:", Yo)

    import numpy as np

    from scipy.stats import pearsonr

    # J1+J2
    x1 = np.array(Xj)
    y1 = np.array(Yj)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z1 = np.polyfit(x1, y1, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p1 = np.poly1d(z1)
    # 在屏幕上打印拟合多项式
    print(p1)

    m1 = p1(x1)
    correlation_coefficient1, p_value = pearsonr(y1, m1)
    print("相关系数：", correlation_coefficient1)

    # 求对应x的各项拟合函数值
    xp1 = np.arange(x1.min(), x1.max(), 1)
    fx1 = p1(xp1)

    # T1+T2
    x2 = np.array(Xt)
    y2 = np.array(Yt)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z2 = np.polyfit(x2, y2, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p2 = np.poly1d(z2)
    # 在屏幕上打印拟合多项式
    print(p2)

    m2 = p2(x2)
    correlation_coefficient2, p_value = pearsonr(y2, m2)
    print("相关系数：", correlation_coefficient2)

    # 求对应x的各项拟合函数值
    xp2 = np.arange(x2.min(), x2.max(), 1)
    fx2 = p2(xp2)

    # R1+R2
    x3 = np.array(Xr)
    y3 = np.array(Yr)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z3 = np.polyfit(x3, y3, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p3 = np.poly1d(z3)
    # 在屏幕上打印拟合多项式
    print(p3)

    m3 = p3(x3)
    correlation_coefficient3, p_value = pearsonr(y3, m3)
    print("相关系数：", correlation_coefficient3)

    # 求对应x的各项拟合函数值
    xp3 = np.arange(x3.min(), x3.max(), 1)
    fx3 = p3(xp3)

    # C1+C2
    x4 = np.array(Xc)
    y4 = np.array(Yc)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z4 = np.polyfit(x4, y4, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p4 = np.poly1d(z4)
    # 在屏幕上打印拟合多项式
    print(p4)

    m4 = p4(x4)
    correlation_coefficient4, p_value = pearsonr(y4, m4)
    print("相关系数：", correlation_coefficient4)

    # 求对应x的各项拟合函数值
    xp4 = np.arange(x4.min(), x4.max(), 1)
    fx4 = p4(xp4)

    # G1+G2
    x5 = np.array(Xg)
    y5 = np.array(Yg)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z5 = np.polyfit(x5, y5, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p5 = np.poly1d(z5)
    # 在屏幕上打印拟合多项式
    print(p5)

    m5 = p5(x5)
    correlation_coefficient5, p_value = pearsonr(y5, m5)
    print("相关系数：", correlation_coefficient5)

    # 求对应x的各项拟合函数值
    xp5 = np.arange(x5.min(), x5.max(), 1)
    fx5 = p5(xp5)

    # O1+O2
    x6 = np.array(Xo)
    y6 = np.array(Yo)
    # 使用polyfit方法来拟合,并选择多项式,这里先使用2次方程
    z6 = np.polyfit(x6, y6, 2)
    # 使用poly1d方法获得多项式系数,按照阶数由高到低排列
    p6 = np.poly1d(z6)
    # 在屏幕上打印拟合多项式
    print(p6)

    m6 = p6(x6)
    correlation_coefficient6, p_value = pearsonr(y6, m6)
    print("相关系数：", correlation_coefficient6)

    # 求对应x的各项拟合函数值
    xp6 = np.arange(x6.min(), x6.max(), 1)
    fx6 = p6(xp6)

    # L1 = LJ-LT
    x = np.arange(x1.min(), x1.max(), 1)
    y = abs(p1(x) - p2(x))
    L1 = y.mean()

    # L2 = LJ-LR
    x = np.arange(x1.min(), x1.max(), 1)
    y = abs(p1(x) - p3(x))
    L2 = y.mean()

    # 寻找left right值
    left = max(x4.min(), x5.min(), x6.min())
    right = min(x4.max(), x5.max(), x6.max())

    print(left)
    print(right)

    # L3 = LJ-LC
    x = np.arange(left, right, 1)
    y = abs(p1(x) - p4(x))
    L3 = y.mean()

    # L4 = LJ-LG
    x = np.arange(left, right, 1)
    y = abs(p1(x) - p5(x))
    L4 = y.mean()

    # L5 = LJ-LO
    x = np.arange(left, right, 1)
    y = abs(p1(x) - p6(x))
    L5 = y.mean()

    # 获取H计算相关数据

    Yj = []
    Yt = []
    Yr = []
    Yc = []
    Yg = []
    Yo = []

    for j in Jlabel:
        totol = []
        for i in data['shapes']:
            if i['label'][0:3] == j:
                totol.append(i['points'][0][1])
        if len(totol) == 2:
            Yj.append(sum(totol) / 2)
        elif len(totol) == 1:
            Yj.append(totol[0])
        else:
            Yj.append(0)
    for j in Tlabel:
        totol = []
        for i in data['shapes']:
            if i['label'][0:3] == j:
                totol.append(i['points'][0][1])
        if len(totol) == 2:
            Yt.append(sum(totol) / 2)
        elif len(totol) == 1:
            Yt.append(totol[0])
        else:
            Yt.append(0)
    for j in Rlabel:
        totol = []
        for i in data['shapes']:
            if i['label'][0:3] == j:
                totol.append(i['points'][0][1])
        if len(totol) == 2:
            Yr.append(sum(totol) / 2)
        elif len(totol) == 1:
            Yr.append(totol[0])
        else:
            Yr.append(0)
    for j in Clabel:
        totol = []
        for i in data['shapes']:
            if i['label'][0:3] == j:
                totol.append(i['points'][0][1])
        if len(totol) == 2:
            Yc.append(sum(totol) / 2)
        elif len(totol) == 1:
            Yc.append(totol[0])
        else:
            Yc.append(0)

    # H1 = LJ-LT
    y = 0
    for i in range(len(Yj)):
        y += abs(Yj[i] - Yt[i])
    H1 = y / len(Yj)

    # H2 = LJ-LR
    y = 0
    for i in range(len(Yj)):
        y += abs(Yj[i] - Yr[i])
    H2 = y / len(Yj)

    # H3 = LJ-LC
    y = 0
    for i in range(len(Yj)):
        y += abs(Yj[i] - Yc[i])
    H3 = y / len(Yj)
    #
    # # H4 = LJ-LG
    # y = 0
    # for i in range(len(Yj)):
    #     y += abs(Yj[i] - Yg[i])
    # H4 = y / len(Yj)
    #
    # # H5 = LJ-LO
    # y = 0
    # for i in range(len(Yj)):
    #     y += abs(Yj[i] - Yo[i])
    # H5 = y / len(Yj)
    #
    print("H1:", H1)
    print("H2:", H2)
    print("H3:", H3)
    # print("H4:", H4)
    # print("H5:", H5)

    print("L1:", L1)
    print("L2:", L2)
    print("L3:", L3)
    print("L4:", L4)
    print("L5:", L5)

    # n0 = L1/L2
    n0 = L1 / L2

    # n1 = L3/L2
    n1 = L3 / L2

    # n2= L4/L2
    n2 = L4 / L2

    # n3=L5/L2
    n3 = L5 / L2

    print("n0:", n0)
    print("n1:", n1)
    print("n2:", n2)
    print("n3:", n3)
    print('')

    import os

    if not os.path.exists("../excel/" + NUM):
        os.mkdir("../excel/" + NUM)

    # with open(NUM + '/d.txt', 'w') as f:
    #     f.write("n0: " + str(n0) + '\n')
    #     f.write("n1: " + str(n1) + '\n')
    #     f.write("n2: " + str(n2) + '\n')
    #     f.write("n3: " + str(n3) + '\n')
    #     f.close()

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill  # 导入填充模块
    wb = load_workbook(filename="../excel/" + NUM + '/' + NUM + '.xlsx')
    wk_sheet = wb['全景牙槽骨吸收']
    Color = ['ffc7ce', '9c0006']  # 红
    fille = PatternFill('solid', fgColor=Color[0])  # 红
    # print(sheet_ranges['E15'].value)
    wk_sheet.cell(row=40, column=12, value=H1)  # 在第12行，第12列下插入“H1”数值
    wk_sheet.cell(row=41, column=12, value=H2)
    wk_sheet.cell(row=42, column=12, value=H3)
    # wk_sheet.cell(row=43, column=12, value=H4)
    # wk_sheet.cell(row=44, column=12, value=H5)

    wk_sheet.cell(row=45, column=12, value=L1)  # 在第17行，第12列下插入“L1”数值
    wk_sheet.cell(row=46, column=12, value=L2)
    wk_sheet.cell(row=47, column=12, value=L3)
    wk_sheet.cell(row=48, column=12, value=L4)
    wk_sheet.cell(row=49, column=12, value=L5)

    # n2＞n1＞n3
    if n2 > n1 > n3:
        wk_sheet.cell(row=50, column=12, value=n0)
        wk_sheet.cell(row=51, column=12, value=n1)
        wk_sheet.cell(row=52, column=12, value=n2)
        wk_sheet.cell(row=53, column=12, value=n3)
    else:
        wk_sheet.cell(row=50, column=12, value=n0).fill = fille
        wk_sheet.cell(row=51, column=12, value=n1).fill = fille
        wk_sheet.cell(row=52, column=12, value=n2).fill = fille
        wk_sheet.cell(row=53, column=12, value=n3).fill = fille

    wk_sheet_1 = wb['拟合程度']
    wk_sheet_1.cell(row=4, column=35, value=correlation_coefficient1)
    wk_sheet_1.cell(row=4, column=36, value=correlation_coefficient2)
    wk_sheet_1.cell(row=4, column=37, value=correlation_coefficient3)
    wk_sheet_1.cell(row=4, column=38, value=correlation_coefficient4)
    wk_sheet_1.cell(row=4, column=39, value=correlation_coefficient5)
    wk_sheet_1.cell(row=4, column=40, value=correlation_coefficient6)

    wb.save("../excel/" + NUM + '/' + NUM + '.xlsx')

    # 绘制图像
    # import matplotlib.pyplot as plt
    # from PIL import Image
    #
    # plt.rc("font", family="Microsoft YaHei")
    # # 插入背景图片
    # img_path = '../img/' + NUM + ".jpg"
    # img = plt.imread(img_path)
    # img_info = Image.open(img_path)
    # fig, ax = plt.subplots()
    # # 指定图片的高度和宽度
    # # ax.imshow(img, extent=[0, 2904, 0, 1536])
    # ax.imshow(img, origin='lower', extent=[0, img_info.width, 0, img_info.height])
    #
    # # 绘制坐标系散点数据及拟合曲线图
    # plot1 = plt.plot(x1, y1, ".", markersize=1, color='blue', label='J')
    # plot3 = plt.plot(x2, y2, '.', markersize=1, color='orange', label='T')
    # plot5 = plt.plot(x3, y3, '.', markersize=1, color='green', label='R')
    # plot7 = plt.plot(x4, y4, '.', markersize=1, color='red', label='C')
    # plot9 = plt.plot(x5, y5, '.', markersize=1, color='m', label='G')
    # plot11 = plt.plot(x6, y6, '.', markersize=1, color='peru', label='O')
    #
    # plot2 = plt.plot(xp1, fx1, 'r', linewidth=0.6, label='polyfit J')
    # plot4 = plt.plot(xp2, fx2, 'b', linewidth=0.6, label='polyfit T')
    # plot6 = plt.plot(xp3, fx3, 'k', linewidth=0.6, label='polyfit R')
    # plot8 = plt.plot(xp4, fx4, 'y', linewidth=0.6, label='polyfit C')
    # plot10 = plt.plot(xp5, fx5, 'c', linewidth=0.6, label='polyfit G')
    # plot12 = plt.plot(xp6, fx6, 'm', linewidth=0.6, label='polyfit O')
    #
    # plt.xlabel('x-price')
    # plt.ylabel('y-amount')
    #
    # # 隐藏刻度值
    # ax = plt.gca()
    # ax.axes.xaxis.set_ticklabels([])
    # ax.axes.yaxis.set_ticklabels([])
    #
    # # 显示图例
    # num1 = 1.01
    # num2 = 0
    # num3 = 3
    # num4 = 0
    # plt.legend(bbox_to_anchor=(num1, num2), loc=num3, borderaxespad=num4)
    # # 右边距调整
    # plt.subplots_adjust(right=0.8)
    # # 坐标轴间距调整
    # plt.xticks(ticks=np.arange(0, img_info.width, 100))
    # plt.yticks(ticks=np.arange(0, img_info.height, 100))
    #
    # if y3.mean() > y2.mean():
    #     # y轴翻转
    #     plt.gca().invert_yaxis()
    #
    # plt.title(NUM)
    # plt.savefig(NUM + '/d.png', dpi=400)
    # plt.show()


# all = ['1-曾悦', '10-宋群英', '11-宋文', '12-唐嘉仪', '13-王柄权', '14-王玮璇', '15-谢川霞', '16-谢容', '17-徐玲玲',
#        '18-杨红梅', '19-李茜颖子', '2-范凯筌', '20-刘娟', '21-刘雪梅', '22-唐平原', '23-陈华', '24-陈静', '25-陈露',
#        '26-郭晓语', '27-黄曦之', '28-敬景怡', '29-邵懿阳', '30-袁杜', '31-赵韫莹', '32-钟明浩', '5-李欣珂',
#        '6-林小婧', '7-罗鸿谕', '8-罗莲', '9-罗雯', '3-韩雨彤']

all = ['31-3', '33-1', '34-5', '42-3', '46-1', '46-7', '46-8', '48-1']

for i in all:
    main(i)