from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # 导入填充模块

files = ['31-3', '33-1', '34-5', '42-3', '46-1', '46-7', '46-8', '48-1']

Color = ['ffc7ce', '9c0006']  # 红
fille = PatternFill('solid', fgColor=Color[0])  # 红

all = load_workbook(filename='../all.xlsx')
sheet_1 = all["全口数据"]
sheet_2 = all["拟合程度"]
i = 4

for file in files:
    print(file)

    data_dict_1 = {
        "name": None,
        "gender": None,
        "age": None,
        "order": None,
        "maxilla": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "mandible": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "a": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "b": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "c": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "d": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "e": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
        "f": {
            "n0": None,
            "n1": None,
            "n2": None,
            "n3": None,
        },
    }

    data_dict_2 = {
        "name": None,
        "gender": None,
        "age": None,
        "order": None,
        "maxilla": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "mandible": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "a": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "b": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "c": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "d": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "e": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
        "f": {
            "J": None,
            "T": None,
            "R": None,
            "C": None,
            "G": None,
            "O": None,
        },
    }

    wb = load_workbook(
        filename='../excel/' + file + '/' + file + '.xlsx')
    wk_sheet_1 = wb['全景牙槽骨吸收']
    wk_sheet_2 = wb['拟合程度']
    data_dict_1['name'] = wk_sheet_1.cell(row=2, column=2).value
    data_dict_1['gender'] = wk_sheet_1.cell(row=3, column=2).value
    data_dict_1['age'] = wk_sheet_1.cell(row=4, column=2).value
    data_dict_1['order'] = wk_sheet_1.cell(row=5, column=2).value

    data_dict_1['maxilla']['n0'] = wk_sheet_1.cell(row=22, column=6).value
    data_dict_1['maxilla']['n1'] = wk_sheet_1.cell(row=23, column=6).value
    data_dict_1['maxilla']['n2'] = wk_sheet_1.cell(row=24, column=6).value
    data_dict_1['maxilla']['n3'] = wk_sheet_1.cell(row=25, column=6).value

    data_dict_1['mandible']['n0'] = wk_sheet_1.cell(row=22, column=12).value
    data_dict_1['mandible']['n1'] = wk_sheet_1.cell(row=23, column=12).value
    data_dict_1['mandible']['n2'] = wk_sheet_1.cell(row=24, column=12).value
    data_dict_1['mandible']['n3'] = wk_sheet_1.cell(row=25, column=12).value

    data_dict_1['a']['n0'] = wk_sheet_1.cell(row=36, column=6).value
    data_dict_1['a']['n1'] = wk_sheet_1.cell(row=37, column=6).value
    data_dict_1['a']['n2'] = wk_sheet_1.cell(row=40, column=6).value
    data_dict_1['a']['n3'] = wk_sheet_1.cell(row=41, column=6).value

    data_dict_1['f']['n0'] = wk_sheet_1.cell(row=36, column=12).value
    data_dict_1['f']['n1'] = wk_sheet_1.cell(row=37, column=12).value
    data_dict_1['f']['n2'] = wk_sheet_1.cell(row=38, column=12).value
    data_dict_1['f']['n3'] = wk_sheet_1.cell(row=39, column=12).value

    data_dict_1['c']['n0'] = wk_sheet_1.cell(row=50, column=6).value
    data_dict_1['c']['n1'] = wk_sheet_1.cell(row=51, column=6).value
    data_dict_1['c']['n2'] = wk_sheet_1.cell(row=52, column=6).value
    data_dict_1['c']['n3'] = wk_sheet_1.cell(row=53, column=6).value

    data_dict_1['d']['n0'] = wk_sheet_1.cell(row=50, column=12).value
    data_dict_1['d']['n1'] = wk_sheet_1.cell(row=51, column=12).value
    data_dict_1['d']['n2'] = wk_sheet_1.cell(row=52, column=12).value
    data_dict_1['d']['n3'] = wk_sheet_1.cell(row=53, column=12).value

    data_dict_1['b']['n0'] = wk_sheet_1.cell(row=64, column=6).value
    data_dict_1['b']['n1'] = wk_sheet_1.cell(row=65, column=6).value
    data_dict_1['b']['n2'] = wk_sheet_1.cell(row=66, column=6).value
    data_dict_1['b']['n3'] = wk_sheet_1.cell(row=67, column=6).value

    data_dict_1['e']['n0'] = wk_sheet_1.cell(row=64, column=12).value
    data_dict_1['e']['n1'] = wk_sheet_1.cell(row=65, column=12).value
    data_dict_1['e']['n2'] = wk_sheet_1.cell(row=66, column=12).value
    data_dict_1['e']['n3'] = wk_sheet_1.cell(row=67, column=12).value

    sheet_1.cell(row=i, column=1, value=str(data_dict_1['order']))
    sheet_1.cell(row=i, column=2, value=str(data_dict_1['name']))
    sheet_1.cell(row=i, column=3, value=str(data_dict_1['gender']))
    sheet_1.cell(row=i, column=4, value=str(data_dict_1['age']))

    sheet_1.cell(row=i, column=5, value=str(data_dict_1['maxilla']['n0']))
    sheet_1.cell(row=i, column=6, value=str(data_dict_1['maxilla']['n1']))
    sheet_1.cell(row=i, column=7, value=str(data_dict_1['maxilla']['n2']))
    sheet_1.cell(row=i, column=8, value=str(data_dict_1['maxilla']['n3']))

    sheet_1.cell(row=i, column=9, value=str(data_dict_1['mandible']['n0']))
    sheet_1.cell(row=i, column=10, value=str(data_dict_1['mandible']['n1']))
    sheet_1.cell(row=i, column=11, value=str(data_dict_1['mandible']['n2']))
    sheet_1.cell(row=i, column=12, value=str(data_dict_1['mandible']['n3']))

    # n2＞n1＞n3
    if data_dict_1['a']['n1'] is not None:
        if data_dict_1['a']['n2'] > data_dict_1['a']['n1'] > data_dict_1['a']['n3']:
            sheet_1.cell(row=i, column=13, value=str(data_dict_1['a']['n0']))
            sheet_1.cell(row=i, column=14, value=str(data_dict_1['a']['n1']))
            sheet_1.cell(row=i, column=15, value=str(data_dict_1['a']['n2']))
            sheet_1.cell(row=i, column=16, value=str(data_dict_1['a']['n3']))
        else:
            sheet_1.cell(row=i, column=13, value=str(data_dict_1['a']['n0'])).fill = fille
            sheet_1.cell(row=i, column=14, value=str(data_dict_1['a']['n1'])).fill = fille
            sheet_1.cell(row=i, column=15, value=str(data_dict_1['a']['n2'])).fill = fille
            sheet_1.cell(row=i, column=16, value=str(data_dict_1['a']['n3'])).fill = fille

    if data_dict_1['b']['n1'] is not None:
        if data_dict_1['b']['n2'] > data_dict_1['b']['n1'] > data_dict_1['b']['n3']:
            sheet_1.cell(row=i, column=17, value=str(data_dict_1['b']['n0']))
            sheet_1.cell(row=i, column=18, value=str(data_dict_1['b']['n1']))
            sheet_1.cell(row=i, column=19, value=str(data_dict_1['b']['n2']))
            sheet_1.cell(row=i, column=20, value=str(data_dict_1['b']['n3']))
        else:
            sheet_1.cell(row=i, column=17, value=str(data_dict_1['b']['n0'])).fill = fille
            sheet_1.cell(row=i, column=18, value=str(data_dict_1['b']['n1'])).fill = fille
            sheet_1.cell(row=i, column=19, value=str(data_dict_1['b']['n2'])).fill = fille
            sheet_1.cell(row=i, column=20, value=str(data_dict_1['b']['n3'])).fill = fille

    if data_dict_1['c']['n1'] is not None:
        if data_dict_1['c']['n2'] > data_dict_1['c']['n1'] > data_dict_1['c']['n3']:
            sheet_1.cell(row=i, column=21, value=str(data_dict_1['c']['n0']))
            sheet_1.cell(row=i, column=22, value=str(data_dict_1['c']['n1']))
            sheet_1.cell(row=i, column=23, value=str(data_dict_1['c']['n2']))
            sheet_1.cell(row=i, column=24, value=str(data_dict_1['c']['n3']))
        else:
            sheet_1.cell(row=i, column=21, value=str(data_dict_1['c']['n0'])).fill = fille
            sheet_1.cell(row=i, column=22, value=str(data_dict_1['c']['n1'])).fill = fille
            sheet_1.cell(row=i, column=23, value=str(data_dict_1['c']['n2'])).fill = fille
            sheet_1.cell(row=i, column=24, value=str(data_dict_1['c']['n3'])).fill = fille

    if data_dict_1['d']['n1'] is not None:
        if data_dict_1['d']['n2'] > data_dict_1['d']['n1'] > data_dict_1['d']['n3']:
            sheet_1.cell(row=i, column=25, value=str(data_dict_1['d']['n0']))
            sheet_1.cell(row=i, column=26, value=str(data_dict_1['d']['n1']))
            sheet_1.cell(row=i, column=27, value=str(data_dict_1['d']['n2']))
            sheet_1.cell(row=i, column=28, value=str(data_dict_1['d']['n3']))
        else:
            sheet_1.cell(row=i, column=25, value=str(data_dict_1['d']['n0'])).fill = fille
            sheet_1.cell(row=i, column=26, value=str(data_dict_1['d']['n1'])).fill = fille
            sheet_1.cell(row=i, column=27, value=str(data_dict_1['d']['n2'])).fill = fille
            sheet_1.cell(row=i, column=28, value=str(data_dict_1['d']['n3'])).fill = fille

    if data_dict_1['e']['n1'] is not None:
        if data_dict_1['e']['n2'] > data_dict_1['e']['n1'] > data_dict_1['e']['n3']:
            sheet_1.cell(row=i, column=29, value=str(data_dict_1['e']['n0']))
            sheet_1.cell(row=i, column=30, value=str(data_dict_1['e']['n1']))
            sheet_1.cell(row=i, column=31, value=str(data_dict_1['e']['n2']))
            sheet_1.cell(row=i, column=32, value=str(data_dict_1['e']['n3']))
        else:
            sheet_1.cell(row=i, column=29, value=str(data_dict_1['e']['n0'])).fill = fille
            sheet_1.cell(row=i, column=30, value=str(data_dict_1['e']['n1'])).fill = fille
            sheet_1.cell(row=i, column=34, value=str(data_dict_1['e']['n2'])).fill = fille
            sheet_1.cell(row=i, column=32, value=str(data_dict_1['e']['n3'])).fill = fille

    if data_dict_1['f']['n1'] is not None:
        if data_dict_1['f']['n2'] > data_dict_1['f']['n1'] > data_dict_1['f']['n3']:
            sheet_1.cell(row=i, column=33, value=str(data_dict_1['f']['n0']))
            sheet_1.cell(row=i, column=34, value=str(data_dict_1['f']['n1']))
            sheet_1.cell(row=i, column=35, value=str(data_dict_1['f']['n2']))
            sheet_1.cell(row=i, column=36, value=str(data_dict_1['f']['n3']))
        else:
            sheet_1.cell(row=i, column=33, value=str(data_dict_1['f']['n0'])).fill = fille
            sheet_1.cell(row=i, column=34, value=str(data_dict_1['f']['n1'])).fill = fille
            sheet_1.cell(row=i, column=35, value=str(data_dict_1['f']['n2'])).fill = fille
            sheet_1.cell(row=i, column=36, value=str(data_dict_1['f']['n3'])).fill = fille

    data_dict_2['order'] = wk_sheet_2.cell(row=4, column=1).value
    data_dict_2['name'] = wk_sheet_2.cell(row=4, column=2).value
    data_dict_2['gender'] = wk_sheet_2.cell(row=4, column=3).value
    data_dict_2['age'] = wk_sheet_2.cell(row=4, column=4).value

    data_dict_2['maxilla']['J'] = wk_sheet_2.cell(row=4, column=5).value
    data_dict_2['maxilla']['T'] = wk_sheet_2.cell(row=4, column=6).value
    data_dict_2['maxilla']['R'] = wk_sheet_2.cell(row=4, column=7).value
    data_dict_2['maxilla']['C'] = wk_sheet_2.cell(row=4, column=8).value
    data_dict_2['maxilla']['G'] = wk_sheet_2.cell(row=4, column=9).value
    data_dict_2['maxilla']['O'] = wk_sheet_2.cell(row=4, column=10).value

    data_dict_2['mandible']['J'] = wk_sheet_2.cell(row=4, column=11).value
    data_dict_2['mandible']['T'] = wk_sheet_2.cell(row=4, column=12).value
    data_dict_2['mandible']['R'] = wk_sheet_2.cell(row=4, column=13).value
    data_dict_2['mandible']['C'] = wk_sheet_2.cell(row=4, column=14).value
    data_dict_2['mandible']['G'] = wk_sheet_2.cell(row=4, column=15).value
    data_dict_2['mandible']['O'] = wk_sheet_2.cell(row=4, column=16).value

    data_dict_2['a']['J'] = wk_sheet_2.cell(row=4, column=17).value
    data_dict_2['a']['T'] = wk_sheet_2.cell(row=4, column=18).value
    data_dict_2['a']['R'] = wk_sheet_2.cell(row=4, column=19).value
    data_dict_2['a']['C'] = wk_sheet_2.cell(row=4, column=20).value
    data_dict_2['a']['G'] = wk_sheet_2.cell(row=4, column=21).value
    data_dict_2['a']['O'] = wk_sheet_2.cell(row=4, column=22).value

    data_dict_2['b']['J'] = wk_sheet_2.cell(row=4, column=23).value
    data_dict_2['b']['T'] = wk_sheet_2.cell(row=4, column=24).value
    data_dict_2['b']['R'] = wk_sheet_2.cell(row=4, column=25).value
    data_dict_2['b']['C'] = wk_sheet_2.cell(row=4, column=26).value
    data_dict_2['b']['G'] = wk_sheet_2.cell(row=4, column=27).value
    data_dict_2['b']['O'] = wk_sheet_2.cell(row=4, column=28).value

    data_dict_2['c']['J'] = wk_sheet_2.cell(row=4, column=29).value
    data_dict_2['c']['T'] = wk_sheet_2.cell(row=4, column=30).value
    data_dict_2['c']['R'] = wk_sheet_2.cell(row=4, column=31).value
    data_dict_2['c']['C'] = wk_sheet_2.cell(row=4, column=32).value
    data_dict_2['c']['G'] = wk_sheet_2.cell(row=4, column=33).value
    data_dict_2['c']['O'] = wk_sheet_2.cell(row=4, column=34).value

    data_dict_2['d']['J'] = wk_sheet_2.cell(row=4, column=35).value
    data_dict_2['d']['T'] = wk_sheet_2.cell(row=4, column=36).value
    data_dict_2['d']['R'] = wk_sheet_2.cell(row=4, column=37).value
    data_dict_2['d']['C'] = wk_sheet_2.cell(row=4, column=38).value
    data_dict_2['d']['G'] = wk_sheet_2.cell(row=4, column=39).value
    data_dict_2['d']['O'] = wk_sheet_2.cell(row=4, column=40).value

    data_dict_2['e']['J'] = wk_sheet_2.cell(row=4, column=41).value
    data_dict_2['e']['T'] = wk_sheet_2.cell(row=4, column=42).value
    data_dict_2['e']['R'] = wk_sheet_2.cell(row=4, column=43).value
    data_dict_2['e']['C'] = wk_sheet_2.cell(row=4, column=44).value
    data_dict_2['e']['G'] = wk_sheet_2.cell(row=4, column=45).value
    data_dict_2['e']['O'] = wk_sheet_2.cell(row=4, column=46).value

    data_dict_2['f']['J'] = wk_sheet_2.cell(row=4, column=47).value
    data_dict_2['f']['T'] = wk_sheet_2.cell(row=4, column=48).value
    data_dict_2['f']['R'] = wk_sheet_2.cell(row=4, column=49).value
    data_dict_2['f']['C'] = wk_sheet_2.cell(row=4, column=50).value
    data_dict_2['f']['G'] = wk_sheet_2.cell(row=4, column=51).value
    data_dict_2['f']['O'] = wk_sheet_2.cell(row=4, column=52).value

    sheet_2.cell(row=i, column=1, value=str(data_dict_2['order']))
    sheet_2.cell(row=i, column=2, value=str(data_dict_2['name']))
    sheet_2.cell(row=i, column=3, value=str(data_dict_2['gender']))
    sheet_2.cell(row=i, column=4, value=str(data_dict_2['age']))

    sheet_2.cell(row=i, column=5, value=str(data_dict_2['maxilla']['J']))
    sheet_2.cell(row=i, column=6, value=str(data_dict_2['maxilla']['T']))
    sheet_2.cell(row=i, column=7, value=str(data_dict_2['maxilla']['R']))
    sheet_2.cell(row=i, column=8, value=str(data_dict_2['maxilla']['C']))
    sheet_2.cell(row=i, column=9, value=str(data_dict_2['maxilla']['G']))
    sheet_2.cell(row=i, column=10, value=str(data_dict_2['maxilla']['O']))

    sheet_2.cell(row=i, column=11, value=str(data_dict_2['mandible']['J']))
    sheet_2.cell(row=i, column=12, value=str(data_dict_2['mandible']['T']))
    sheet_2.cell(row=i, column=13, value=str(data_dict_2['mandible']['R']))
    sheet_2.cell(row=i, column=14, value=str(data_dict_2['mandible']['C']))
    sheet_2.cell(row=i, column=15, value=str(data_dict_2['mandible']['G']))
    sheet_2.cell(row=i, column=16, value=str(data_dict_2['mandible']['O']))

    sheet_2.cell(row=i, column=17, value=str(data_dict_2['a']['J']))
    sheet_2.cell(row=i, column=18, value=str(data_dict_2['a']['T']))
    sheet_2.cell(row=i, column=19, value=str(data_dict_2['a']['R']))
    sheet_2.cell(row=i, column=20, value=str(data_dict_2['a']['C']))
    sheet_2.cell(row=i, column=21, value=str(data_dict_2['a']['G']))
    sheet_2.cell(row=i, column=22, value=str(data_dict_2['a']['O']))

    sheet_2.cell(row=i, column=23, value=str(data_dict_2['b']['J']))
    sheet_2.cell(row=i, column=24, value=str(data_dict_2['b']['T']))
    sheet_2.cell(row=i, column=25, value=str(data_dict_2['b']['R']))
    sheet_2.cell(row=i, column=26, value=str(data_dict_2['b']['C']))
    sheet_2.cell(row=i, column=27, value=str(data_dict_2['b']['G']))
    sheet_2.cell(row=i, column=28, value=str(data_dict_2['b']['O']))

    sheet_2.cell(row=i, column=29, value=str(data_dict_2['c']['J']))
    sheet_2.cell(row=i, column=30, value=str(data_dict_2['c']['T']))
    sheet_2.cell(row=i, column=31, value=str(data_dict_2['c']['R']))
    sheet_2.cell(row=i, column=32, value=str(data_dict_2['c']['C']))
    sheet_2.cell(row=i, column=33, value=str(data_dict_2['c']['G']))
    sheet_2.cell(row=i, column=34, value=str(data_dict_2['c']['O']))

    sheet_2.cell(row=i, column=35, value=str(data_dict_2['d']['J']))
    sheet_2.cell(row=i, column=36, value=str(data_dict_2['d']['T']))
    sheet_2.cell(row=i, column=37, value=str(data_dict_2['d']['R']))
    sheet_2.cell(row=i, column=38, value=str(data_dict_2['d']['C']))
    sheet_2.cell(row=i, column=39, value=str(data_dict_2['d']['G']))
    sheet_2.cell(row=i, column=40, value=str(data_dict_2['d']['O']))

    sheet_2.cell(row=i, column=41, value=str(data_dict_2['e']['J']))
    sheet_2.cell(row=i, column=42, value=str(data_dict_2['e']['T']))
    sheet_2.cell(row=i, column=43, value=str(data_dict_2['e']['R']))
    sheet_2.cell(row=i, column=44, value=str(data_dict_2['e']['C']))
    sheet_2.cell(row=i, column=45, value=str(data_dict_2['e']['G']))
    sheet_2.cell(row=i, column=46, value=str(data_dict_2['e']['O']))

    sheet_2.cell(row=i, column=47, value=str(data_dict_2['f']['J']))
    sheet_2.cell(row=i, column=48, value=str(data_dict_2['f']['T']))
    sheet_2.cell(row=i, column=49, value=str(data_dict_2['f']['R']))
    sheet_2.cell(row=i, column=50, value=str(data_dict_2['f']['C']))
    sheet_2.cell(row=i, column=51, value=str(data_dict_2['f']['G']))
    sheet_2.cell(row=i, column=52, value=str(data_dict_2['f']['O']))

    i += 1
    # print(data_dict_1)

all.save('../normal.xlsx')
